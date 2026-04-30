"""B3 "Movimentação" XLSX extractor.

Parses the official B3 *Movimentação* report exported from
https://www.investidor.b3.com.br/ (menu Extratos → Movimentação → Exportar Excel).

Scope (V1):
* Only **provento** rows are emitted: ``Dividendo``, ``Juros Sobre Capital
  Próprio`` and ``Rendimento``. Everything else (trades, transfers, securities
  lending, corporate events, fixed-income flows, contábil "- Transferido"
  duplicates, etc.) is silently ignored — those concerns are handled by other
  extractors or are out of scope.

Field mapping per provento row:
    Quantidade × Preço unitário  → ``gross_value`` (announced gross per share × qty)
    Valor da Operação            → ``net_value`` (what actually hit the account)
    fees = gross − net           → IR retained at source (only > 0 for JCP)

The expected XLSX header (case-insensitive):
    Entrada/Saída | Data | Movimentação | Produto | Instituição |
    Quantidade | Preço unitário | Valor da Operação
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from extractors.base import BaseExtractor, ExtractionResult

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

#: B3 movement names that should be imported, mapped to canonical operation
#: types understood by ``normalizers.validator.normalise_operation_type``.
_PROVENT_MAP: dict[str, str] = {
    "dividendo": "dividend",
    "juros sobre capital próprio": "jcp",
    "juros sobre capital proprio": "jcp",
    "rendimento": "rendimento",
}

#: Required columns to recognise the file as a B3 Movimentação export.
_REQUIRED_HEADERS: set[str] = {
    "entrada/saída",
    "data",
    "movimentação",
    "produto",
    "instituição",
    "quantidade",
    "preço unitário",
    "valor da operação",
}


def _normalise_header(value: Any) -> str:
    return str(value or "").strip().lower()


def _parse_ticker(produto: Any) -> str | None:
    """Extract the ticker from a ``Produto`` cell.

    The B3 format is ``"TICKER - DESCRIÇÃO COMPLETA"``. For a few CDB/Tesouro
    products the format differs, but we never reach this code path for those
    movement types in V1.
    """
    raw = str(produto or "").strip()
    if not raw:
        return None
    # Split on the first " - " (with surrounding spaces) and take the head.
    ticker = re.split(r"\s+-\s+", raw, maxsplit=1)[0].strip().upper()
    return ticker or None


def _normalise_institution(value: Any) -> str:
    """Compact the institution name to the first significant token.

    Brokers vary in how they report the same institution across months
    (``"INTER DTVM"`` vs ``"INTER DISTRIBUIDORA DE TÍTULOS..."``). To keep the
    deterministic ``external_id`` stable across these variations we collapse to
    the first non-generic token, uppercased.
    """
    raw = str(value or "").strip().upper()
    if not raw:
        return "unknown"
    # Drop common suffixes that don't add identity.
    raw = raw.split(" - ", 1)[0]
    first = raw.split()[0] if raw.split() else raw
    return first or "unknown"


def _parse_date(value: Any) -> str | None:
    """Convert B3 dates (``DD/MM/AAAA`` or native datetime) to ISO 8601."""
    if value is None:
        return None
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    raw = str(value).strip()
    if not raw:
        return None
    # B3 always uses DD/MM/AAAA in the XLSX export.
    parts = raw.split("/")
    if len(parts) == 3:
        d, m, y = parts
        try:
            return f"{int(y):04d}-{int(m):02d}-{int(d):02d}"
        except ValueError:
            return None
    return None


def _parse_number(value: Any) -> float | None:
    """Parse a numeric cell. Returns None for ``-`` placeholders or empty."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    raw = str(value).strip()
    if not raw or raw == "-":
        return None
    raw = raw.replace("R$", "").replace("\xa0", " ").strip()
    if "," in raw and "." in raw:
        raw = raw.replace(".", "").replace(",", ".")
    elif "," in raw:
        raw = raw.replace(",", ".")
    try:
        return float(raw)
    except ValueError:
        return None


def _to_cents(value: float | None) -> int:
    if value is None:
        return 0
    return int(round(value * 100))


# ---------------------------------------------------------------------------
# Extractor
# ---------------------------------------------------------------------------

class B3MovimentacaoXlsxExtractor(BaseExtractor):
    """Extracts provento rows from a B3 Movimentação XLSX export."""

    source_type = "b3_movimentacao_xlsx"

    EXTRACTOR_VERSION = 1

    def can_handle(self, file_path: Path) -> bool:
        if file_path.suffix.lower() != ".xlsx":
            return False
        try:
            import openpyxl  # noqa: PLC0415

            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            headers = {
                _normalise_header(cell.value)
                for cell in next(ws.iter_rows(min_row=1, max_row=1))
                if cell.value
            }
            wb.close()
        except Exception:  # noqa: BLE001
            return False
        return _REQUIRED_HEADERS.issubset(headers)

    def extract(self, file_path: Path) -> ExtractionResult:
        result = ExtractionResult(source_type=self.source_type)
        try:
            import openpyxl  # noqa: PLC0415

            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as exc:  # noqa: BLE001
            result.errors.append(
                {
                    "row_index": None,
                    "error_type": "parsing",
                    "message": f"Could not read B3 Movimentação file: {exc}",
                }
            )
            return result

        if not rows:
            return result

        # Map header → column index.
        header_idx: dict[str, int] = {
            _normalise_header(cell): idx for idx, cell in enumerate(rows[0])
        }
        missing = _REQUIRED_HEADERS - header_idx.keys()
        if missing:
            result.errors.append(
                {
                    "row_index": 0,
                    "error_type": "parsing",
                    "message": f"Missing required headers: {sorted(missing)}",
                }
            )
            return result

        col_movement = header_idx["movimentação"]
        col_entrada = header_idx["entrada/saída"]
        col_date = header_idx["data"]
        col_produto = header_idx["produto"]
        col_inst = header_idx["instituição"]
        col_qty = header_idx["quantidade"]
        col_pu = header_idx["preço unitário"]
        col_valor = header_idx["valor da operação"]

        for row_index, row in enumerate(rows[1:], start=2):
            if not row or not any(c is not None and str(c).strip() != "" for c in row):
                continue

            movement_raw = str(row[col_movement] or "").strip()
            movement_key = movement_raw.lower()

            # V1 only emits proventos; everything else is intentionally skipped.
            operation_type = _PROVENT_MAP.get(movement_key)
            if operation_type is None:
                continue

            # Defensive: only count credits. The B3 file never reports debit
            # for these three types in practice, but if it ever does we want
            # to skip rather than negate a payment.
            if str(row[col_entrada] or "").strip().lower() not in {"credito", "crédito"}:
                continue

            operation_date = _parse_date(row[col_date])
            ticker = _parse_ticker(row[col_produto])
            qty = _parse_number(row[col_qty])
            unit_price = _parse_number(row[col_pu])
            gross_or_net = _parse_number(row[col_valor])

            if not operation_date or not ticker or qty is None or gross_or_net is None:
                result.errors.append(
                    {
                        "row_index": row_index,
                        "error_type": "validation",
                        "message": (
                            f"Provento row missing required fields "
                            f"(date={row[col_date]!r}, produto={row[col_produto]!r}, "
                            f"qty={row[col_qty]!r}, valor={row[col_valor]!r})"
                        ),
                    }
                )
                continue

            net_cents = _to_cents(gross_or_net)
            # Compute gross from qty × pu when the unit-price column is filled.
            # Otherwise (rare for proventos) fall back to the net value.
            gross_cents = (
                _to_cents(qty * unit_price) if unit_price is not None else net_cents
            )
            # Sanity: gross must never be smaller than net for proventos.
            # If the file ever swaps the convention, fall back to net for both.
            if gross_cents < net_cents:
                gross_cents = net_cents
            ir_cents = max(0, gross_cents - net_cents)

            institution = _normalise_institution(row[col_inst])
            external_id = (
                f"b3mov:{operation_date}:{operation_type}:{ticker}:"
                f"{institution}:{net_cents}"
            )

            # Monetary fields are passed in BRL (reais) — the normalizer
            # (parse_monetary_cents) is responsible for converting to cents.
            gross_reais = gross_cents / 100.0
            ir_reais = ir_cents / 100.0
            unit_price_reais = float(unit_price) if unit_price is not None else 0.0

            record: dict[str, Any] = {
                "source": self.source_type,
                "external_id": external_id,
                "asset_code": ticker,
                # asset_type left empty so the normalizer infers it from the ticker.
                "operation_type": operation_type,
                "operation_date": operation_date,
                "quantity": qty,
                "unit_price": unit_price_reais,
                "gross_value": gross_reais,
                "fees": ir_reais,  # IR retido (only > 0 for JCP in practice)
                "broker": str(row[col_inst] or "").strip() or None,
                "notes": movement_raw,  # preserve original wording for audit
                "file_name": file_path.name,
            }
            result.records.append(record)

        return result


__all__ = ["B3MovimentacaoXlsxExtractor"]
