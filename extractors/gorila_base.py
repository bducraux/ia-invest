"""Shared Gorila XLSX extraction primitives.

Gorila exports across portfolios share a common tabular shape with columns like:
    Ativo | Tipo | Data da transação | Quantidade | Preço | Valor total | Custodiante

Some exports also include an extra leading ``Data de modificação`` column.
Portfolio-specific subclasses configure the source type, default asset type,
operation mapping, and any asset-code aliases.
"""

from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from extractors.base import BaseExtractor, ExtractionResult


def excel_serial_to_date(serial: Any) -> str | None:
    """Convert Excel date serial number or date string to ISO 8601 string (YYYY-MM-DD)."""
    if serial is None:
        return None
    # Already a datetime/date object
    if hasattr(serial, "strftime"):
        return serial.strftime("%Y-%m-%d")
    # String date: try common formats
    if isinstance(serial, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(serial.strip(), fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return None
    # Excel serial number
    try:
        n = int(serial)
        dt = datetime(1899, 12, 30) + timedelta(days=n)
        return dt.strftime("%Y-%m-%d")
    except (ValueError, TypeError, OverflowError):
        return None


def parse_brl(value: Any) -> float:
    """Parse Brazilian-locale currency string to float."""
    if value is None:
        return 0.0
    s = str(value).replace("R$", "").replace("\xa0", " ").strip()
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0


def parse_quantity_value(value: Any) -> float:
    """Parse quantity values from numeric cells or locale-formatted strings."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    raw = str(value).replace("\xa0", " ").strip()
    if not raw:
        return 0.0

    if "," in raw and "." in raw:
        raw = raw.replace(".", "").replace(",", ".")
    elif "," in raw:
        raw = raw.replace(",", ".")

    return float(raw)


class BaseGorilaXlsxExtractor(BaseExtractor):
    """Base extractor for Gorila XLSX portfolio exports."""

    source_type = ""
    external_id_prefix = ""
    default_asset_type: str | None = None
    operation_type_map: dict[str, str] = {
        "compra": "buy",
        "venda": "sell",
    }
    asset_code_aliases: dict[str, str] = {}

    _EXPECTED_HEADERS = {
        "data de modificação",
        "data de modificacao",
        "ativo",
        "tipo",
        "data da transação",
        "data da transacao",
        "quantidade",
        "preço",
        "preco",
        "valor total",
        "custodiante",
    }

    def can_handle(self, file_path: Path) -> bool:
        if file_path.suffix.lower() != ".xlsx":
            return False
        try:
            import openpyxl  # noqa: PLC0415

            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            headers = {
                str(cell.value).strip().lower()
                for cell in next(ws.iter_rows(min_row=1, max_row=1))
                if cell.value
            }
            wb.close()
            return {"ativo", "tipo", "data da transação", "quantidade", "valor total"}.issubset(headers) or {
                "ativo", "tipo", "data da transacao", "quantidade", "valor total"
            }.issubset(headers)
        except Exception:  # noqa: BLE001
            return False

    def extract(self, file_path: Path) -> ExtractionResult:
        result = ExtractionResult(source_type=self.source_type)

        try:
            import openpyxl  # noqa: PLC0415

            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
        except Exception as exc:  # noqa: BLE001
            result.errors.append({"row": 0, "error": f"Cannot open file: {exc}"})
            return result

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            result.errors.append({"row": 0, "error": "File is empty"})
            return result

        raw_headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        col: dict[str, int] = {}
        for i, h in enumerate(raw_headers):
            if "ativo" in h:
                col["asset_code"] = i
            elif h == "tipo":
                col["operation_type"] = i
            elif "data da transação" in h or "data da transacao" in h:
                col["operation_date"] = i
            elif "data de modificação" in h or "data de modificacao" in h:
                col["modification_date"] = i
            elif "quantidade" in h:
                col["quantity"] = i
            elif "preço" in h or "preco" in h:
                col["unit_price"] = i
            elif "valor total" in h:
                col["gross_value"] = i
            elif "custodiante" in h:
                col["broker"] = i

        required = {"asset_code", "operation_type", "operation_date", "quantity", "gross_value"}
        missing = required - col.keys()
        if missing:
            result.errors.append(
                {
                    "row": 1,
                    "error": f"Missing required columns: {missing}. Found: {raw_headers}",
                }
            )
            return result

        for row_idx, row in enumerate(rows[1:], start=2):
            if not any(row):
                continue

            def get(key: str) -> Any:
                idx = col.get(key)
                return row[idx] if idx is not None and idx < len(row) else None

            raw_asset = str(get("asset_code") or "").strip().upper()
            raw_type = str(get("operation_type") or "").strip().lower()
            raw_date = get("operation_date")
            raw_qty = get("quantity")
            raw_price = get("unit_price")
            raw_gross = get("gross_value")
            raw_broker = str(get("broker") or "").strip()

            if not raw_asset:
                result.errors.append({"row": row_idx, "error": "Missing asset code"})
                continue

            operation_type = self.operation_type_map.get(raw_type)
            if not operation_type:
                result.errors.append(
                    {
                        "row": row_idx,
                        "error": f"Unknown operation type: {get('operation_type')!r}",
                    }
                )
                continue

            operation_date = excel_serial_to_date(raw_date)
            if not operation_date:
                result.errors.append(
                    {
                        "row": row_idx,
                        "error": f"Cannot parse date: {raw_date!r}",
                    }
                )
                continue

            try:
                quantity = parse_quantity_value(raw_qty)
            except ValueError:
                result.errors.append(
                    {"row": row_idx, "error": f"Cannot parse quantity: {raw_qty!r}"}
                )
                continue

            unit_price = parse_brl(raw_price)
            gross_value = parse_brl(raw_gross)
            asset_code = self.asset_code_aliases.get(raw_asset, raw_asset)
            external_id_prefix = self.external_id_prefix or self.source_type

            record = {
                "source": self.source_type,
                "external_id": f"{external_id_prefix}:{operation_date}:{asset_code}:{operation_type}:{quantity:.8f}",
                "asset_code": asset_code,
                "asset_type": self.default_asset_type,
                "operation_type": operation_type,
                "operation_date": operation_date,
                "quantity": quantity,
                "unit_price": unit_price,
                "gross_value": gross_value,
                "fees": 0,
                "broker": raw_broker,
                "file_name": file_path.name,
            }
            self._enrich_record(record, row, col)
            result.records.append(record)

        return result

    def _enrich_record(self, record: dict[str, Any], row: tuple[Any, ...], col: dict[str, int]) -> None:
        """Hook for subclasses to add extra fields to the raw record."""
